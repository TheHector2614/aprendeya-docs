import openpyxl
from pathlib import Path

sh = Path("raw") / "sharepoint"
dh = Path("raw") / "drive"

# ═══ FIN-003 — Estados Financieros 2025 ═══
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Balance General 2025"
ws.append(["Cuenta", "Dic 2025", "Dic 2024"])
for r in [
    ["Activo Corriente", 850000000, 620000000],
    ["Efectivo", 320000000, 210000000],
    ["Cuentas por Cobrar", 380000000, 290000000],
    ["Inventarios", 150000000, 120000000],
    ["Activo No Corriente", 1200000000, 950000000],
    ["TOTAL ACTIVO", 2050000000, 1570000000],
    ["Pasivo Corriente", 520000000, 410000000],
    ["Pasivo No Corriente", 380000000, 290000000],
    ["TOTAL PASIVO", 900000000, 700000000],
    ["Patrimonio", 1150000000, 870000000],
    ["TOTAL PATRIMONIO", 1150000000, 870000000],
]:
    ws.append(r)
ws2 = wb.create_sheet("P&L 2025")
ws2.append(["Cuenta", "2025", "2024"])
for r in [
    ["Ingresos Operacionales", 3200000000, 2100000000],
    ["Utilidad Bruta", 1800000000, 1150000000],
    ["Utilidad Operativa", 600000000, 300000000],
    ["Gastos Financieros", -40000000, -30000000],
    ["Utilidad Neta", 560000000, 270000000],
]:
    ws2.append(r)
ws3 = wb.create_sheet("Resumen Financiero")
ws3.append(["Tema", "Explicacion"])
ws3.append(["Ingresos totales 2025", "Los ingresos operacionales de AprendeYa en 2025 fueron de $3,200 millones de pesos colombianos, compuestos por $1,800 millones en matriculas de cursos individuales, $950 millones en suscripciones premium y $450 millones en cursos corporativos para empresas. Esto representa un crecimiento del 52.4% frente a los $2,100 millones de 2024."])
ws3.append(["Utilidad neta 2025", "La utilidad neta de la empresa en 2025 fue de $560 millones de pesos, mas del doble que los $270 millones obtenidos en 2024. El margen neto mejoro del 12.9% al 17.5%, reflejando mayor eficiencia operativa y economias de escala."])
ws3.append(["Balance general", "Al cierre de 2025, AprendeYa tiene activos totales por $2,050 millones (incluyendo $850 millones en activos corrientes y $1,200 millones en activos no corrientes como propiedad, equipo e intangibles), pasivos totales por $900 millones y un patrimonio neto de $1,150 millones. La empresa no tiene deuda financiera significativa."])
ws3.append(["Crecimiento financiero", "AprendeYa duplico su utilidad neta en 2025 impulsada por el crecimiento en matriculas de cursos de programacion e idiomas. La empresa es rentable y genera flujo de caja positivo."])
wb.save(str(sh / "fin" / "FIN-003.xlsx"))
print("FIN-003.xlsx OK")

# ═══ FIN-005 — Reembolso de Gastos ═══
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Politica de Gastos"
ws.append(["Concepto", "Tope Diario", "Requiere Autorizacion", "Soporte"])
for r in [
    ["Alimentacion", 45000, "No", "Factura"],
    ["Transporte terrestre", 80000, "No", "Factura / tiquete"],
    ["Transporte aereo nacional", "", "Si - Lider", "Tiquete + boarding"],
    ["Hospedaje", 250000, "Si - Lider", "Factura"],
    ["Viaticos internacionales", "USD 120/dia", "Si - Direccion", "Factura + itinerario"],
    ["Gasolina (vehiculo propio)", "$1,800/km", "No", "Registro de kilometraje"],
    ["Peajes y parqueaderos", "Costo real", "No", "Factura"],
    ["Papeleria y suministros", 50000, "No", "Factura"],
    ["Suscripciones herramientas", "", "Si - TI", "Factura"],
]:
    ws.append(r)
ws2 = wb.create_sheet("Procedimiento")
ws2.append(["Paso", "Descripcion", "Plazo"])
ws2.append(["1", "Realizar el gasto con recursos propios", ""])
ws2.append(["2", "Solicitar factura a nombre de AprendeYa", ""])
ws2.append(["3", "Ingresar al sistema de gastos y subir la factura", "15 dias"])
ws2.append(["4", "El lider directo aprueba el gasto", "3 dias habiles"])
ws2.append(["5", "Finanzas procesa el reembolso", "5 dias habiles"])
ws2.append(["6", "El valor se acredita en la nomina siguiente", ""])
ws3 = wb.create_sheet("Resumen Gastos")
ws3.append(["Tema", "Explicacion"])
ws3.append(["Tope alimentacion diario", "El gasto maximo diario para alimentacion en viajes de trabajo es de $45,000 pesos colombianos. No requiere autorizacion previa del lider. Se debe presentar la factura o soporte de pago. Aplica para desayuno, almuerzo y cena."])
ws3.append(["Tope transporte diario", "El transporte terrestre (taxi, bus, uber) tiene un tope de $80,000 pesos diarios. No necesita autorizacion. Los tiquetes aereos nacionales requieren autorizacion del lider directo antes de la compra."])
ws3.append(["Tope hospedaje", "El alojamiento en hotel tiene un maximo de $250,000 pesos por noche. Requiere autorizacion del lider directo. Se debe presentar la factura del hotel con los datos fiscales de AprendeYa."])
ws3.append(["Viaticos internacionales", "Para viajes al exterior, el viatico diario es de USD 120 dolares que cubre alimentacion y gastos menores. Requiere autorizacion de la direccion. Se debe presentar factura mas itinerario de viaje."])
ws3.append(["Reembolso gasolina vehiculo propio", "Si el colaborador usa su vehiculo particular para fines laborales, se reembolsa a $1,800 pesos por kilometro recorrido. Se debe llevar un registro de kilometraje con origen, destino y motivo del viaje."])
ws3.append(["Procedimiento para reembolso", "El colaborador paga el gasto de su propio bolsillo. Luego sube la factura al sistema de gastos dentro de los 15 dias siguientes al gasto. El lider directo aprueba en un maximo de 3 dias habiles. Finanzas procesa el reembolso en 5 dias habiles y el valor se acredita en la siguiente quincena de nomina."])
wb.save(str(sh / "fin" / "FIN-005.xlsx"))
print("FIN-005.xlsx OK")

# ═══ FIN-004 — Proyeccion Financiera ═══
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proyeccion 2026-2027"
ws.append(["Indicador", "2025 Real", "2026 Proy", "2027 Proy"])
for r in [
    ["Ingresos", 3200000000, 4500000000, 5800000000],
    ["Estudiantes Pagos", 8500, 15000, 22000],
    ["CAC (USD)", 25, 20, 18],
    ["LTV (USD)", 150, 200, 250],
    ["Margen Neto (%)", 18, 22, 25],
    ["NPS", 45, 55, 65],
    ["Cursos en Catalogo", 40, 80, 120],
]:
    ws.append(r)
ws2 = wb.create_sheet("Supuestos")
ws2.append(["Variable", "Valor"])
ws2.append(["Tasa de conversion", "4.5%"])
ws2.append(["Retencion mensual", "88%"])
ws2.append(["Renovacion anual premium", "75%"])
ws3 = wb.create_sheet("Resumen Proyeccion")
ws3.append(["Indicador", "Detalle"])
ws3.append(["Proyeccion ingresos 2026", "Se proyectan ingresos por $4,500 millones en 2026, un crecimiento del 41% respecto a 2025, alcanzando $5,800 millones en 2027."])
ws3.append(["Proyeccion estudiantes", "Se espera pasar de 8,500 estudiantes pagos en 2025 a 15,000 en 2026 y 22,000 en 2027."])
ws3.append(["CAC y LTV", "El costo de adquisicion de clientes (CAC) se reducira de $25 USD a $20 USD. El valor de vida del cliente (LTV) subira de $150 a $200 USD."])
ws3.append(["Rentabilidad proyectada", "El margen neto pasara del 18% actual al 22% en 2026 y 25% en 2027 por economias de escala."])
ws3.append(["Meta de NPS", "El NPS pasara de 45 en 2025 a 55 en 2026 y 65 en 2027."])
wb.save(str(sh / "fin" / "FIN-004.xlsx"))
print("FIN-004.xlsx OK")

print("\nXLSX mejorados OK")
