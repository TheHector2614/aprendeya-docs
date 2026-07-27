import openpyxl, docx
from pathlib import Path

BASE = Path("raw")
sh = BASE / "sharepoint"
dh = BASE / "drive"

# ════════════════════════════════════════════════
# FIN-003.xlsx — Estados Financieros 2025
# ════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Balance General 2025"
ws.append(["Cuenta", "Dic 2025", "Dic 2024"])
for r in [
    ["Activo Corriente", 850000000, 620000000],
    ["Efectivo", 320000000, 210000000],
    ["Cuentas por Cobrar", 380000000, 290000000],
    ["Inventarios", 150000000, 120000000],
    ["Activo No Corriente", 1200000000, 950000000],
    ["Propiedad y Equipo", 780000000, 610000000],
    ["Intangibles", 420000000, 340000000],
    ["TOTAL ACTIVO", 2050000000, 1570000000],
    ["", "", ""],
    ["Pasivo Corriente", 520000000, 410000000],
    ["Cuentas por Pagar", 280000000, 230000000],
    ["Obligaciones Laborales", 180000000, 140000000],
    ["Impuestos por Pagar", 60000000, 40000000],
    ["Pasivo No Corriente", 380000000, 290000000],
    ["TOTAL PASIVO", 900000000, 700000000],
    ["", "", ""],
    ["Patrimonio", 1150000000, 870000000],
    ["Capital Social", 500000000, 500000000],
    ["Utilidades Retenidas", 650000000, 370000000],
    ["TOTAL PATRIMONIO", 1150000000, 870000000],
]:
    ws.append(r)
ws2 = wb.create_sheet("P&L 2025")
ws2.append(["Cuenta", "2025", "2024"])
for r in [
    ["Ingresos Operacionales", 3200000000, 2100000000],
    ["Matriculas", 1800000000, 1200000000],
    ["Suscripciones", 950000000, 600000000],
    ["Cursos Corporativos", 450000000, 300000000],
    ["Costo de Ventas", -1400000000, -950000000],
    ["Utilidad Bruta", 1800000000, 1150000000],
    ["Gastos Operativos", -1200000000, -850000000],
    ["Tecnologia", -350000000, -250000000],
    ["Marketing", -280000000, -200000000],
    ["Nomina", -480000000, -320000000],
    ["Generales", -90000000, -80000000],
    ["Utilidad Operativa", 600000000, 300000000],
    ["Gastos Financieros", -40000000, -30000000],
    ["Utilidad Neta", 560000000, 270000000],
]:
    ws2.append(r)

# Hoja de texto descriptivo para mejor busqueda semantica
ws3 = wb.create_sheet("Resumen Financiero")
ws3.append(["Seccion", "Descripcion"])
ws3.append(["Ingresos 2025", "Los ingresos operacionales de AprendeYa en 2025 fueron de $3,200 millones de pesos, compuestos por $1,800 millones en matriculas, $950 millones en suscripciones y $450 millones en cursos corporativos. Esto representa un crecimiento del 52% frente a los $2,100 millones de 2024."])
ws3.append(["Utilidad Neta 2025", "La utilidad neta de 2025 fue de $560 millones de pesos, mas del doble de los $270 millones obtenidos en 2024. El margen neto paso de 13% a 18%."])
ws3.append(["Balance General", "Al cierre de 2025, AprendeYa tiene activos totales por $2,050 millones ($850 millones corrientes y $1,200 millones no corrientes), pasivos por $900 millones y un patrimonio de $1,150 millones."])
ws3.append(["Crecimiento", "La empresa duplico su utilidad neta en 2025 respecto a 2024, pasando de $270 millones a $560 millones, impulsada por el crecimiento en matriculas y suscripciones."])
wb.save(str(sh / "fin" / "FIN-003.xlsx"))
print("FIN-003.xlsx OK (con hoja Resumen)")

# ════════════════════════════════════════════════
# FIN-004.xlsx — Proyeccion Financiera 2026-2027
# ════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Proyeccion 2026-2027"
ws.append(["Indicador", "2025 Real", "2026 Proy", "2027 Proy"])
for r in [
    ["Ingresos", 3200000000, 4500000000, 5800000000],
    ["Crecimiento (%)", "", 41, 29],
    ["Estudiantes Pagos", 8500, 15000, 22000],
    ["Ticket Promedio", 376000, 300000, 264000],
    ["CAC (USD)", 25, 20, 18],
    ["LTV (USD)", 150, 200, 250],
    ["Margen Bruto (%)", 56, 60, 62],
    ["Margen Neto (%)", 18, 22, 25],
    ["NPS", 45, 55, 65],
    ["Colaboradores", 45, 65, 85],
    ["Cursos en Catalogo", 40, 80, 120],
]:
    ws.append(r)
ws2 = wb.create_sheet("Supuestos")
ws2.append(["Variable", "Valor"])
ws2.append(["Tasa de conversion lead-estudiante", "4.5%"])
ws2.append(["Tasa de retencion mensual", "88%"])
ws2.append(["Renovacion anual planes premium", "75%"])
ws2.append(["Crecimiento catalogo anual", "100%"])
ws2.append(["Inversion en tecnologia (% ingresos)", "12%"])
ws2.append(["Inversion en marketing (% ingresos)", "10%"])
ws3 = wb.create_sheet("Resumen Proyeccion")
ws3.append(["Indicador", "Valor"])
ws3.append(["Ingresos proyectados 2026", "Se proyectan ingresos por $4,500 millones en 2026, un crecimiento del 41% respecto a 2025, alcanzando $5,800 millones en 2027."])
ws3.append(["Estudiantes", "Se espera pasar de 8,500 estudiantes pagos en 2025 a 15,000 en 2026 y 22,000 en 2027."])
ws3.append(["Eficiencia", "El CAC se reducira de $25 USD a $20 USD en 2026, mientras el LTV aumentara de $150 a $200 USD. La relacion LTV/CAC pasara de 6x a 10x."])
ws3.append(["Rentabilidad", "El margen neto proyectado es de 22% en 2026 y 25% en 2027, reflejando economias de escala."])
ws3.append(["NPS", "La meta de NPS es 55 en 2026 y 65 en 2027, mejorando desde el 45 actual."])
wb.save(str(sh / "fin" / "FIN-004.xlsx"))
print("FIN-004.xlsx OK (con hoja Resumen)")

# ════════════════════════════════════════════════
# FIN-005.xlsx — Reembolso de Gastos y Viaticos
# ════════════════════════════════════════════════
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Politica de Gastos"
ws.append(["Concepto", "Tope Diario", "Requiere Autorizacion", "Soporte"])
for r in [
    ["Alimentacion", 45000, "No", "Factura"],
    ["Transporte terrestre", 80000, "No", "Factura / tiquete"],
    ["Transporte aereo nacional", "", "Si - Lider", "Tiquete + boarding"],
    ["Hospedaje", 250000, "Si - Lider", "Factura"],
    ["Viaticos internacionales", "USD 120/dia", "Si - Direccion", "Factura + itinerario"],
    ["Gasolina (vehiculo propio)", "$1,800/km", "No", "Registro de kilometraje"],
    ["Peajes y parqueaderos", "Costo real", "No", "Factura"],
    ["Papeleria y suministros", 50000, "No", "Factura"],
    ["Suscripciones herramientas", "", "Si - TI", "Factura"],
]:
    ws.append(r)
ws2 = wb.create_sheet("Procedimiento")
ws2.append(["Paso", "Descripcion", "Plazo"])
ws2.append(["1", "Realizar el gasto con recursos propios", ""])
ws2.append(["2", "Solicitar factura a nombre de AprendeYa (NIT 901.XXX.XXX)", ""])
ws2.append(["3", "Ingresar al sistema de gastos y subir la factura", "Dentro de 15 dias del gasto"])
ws2.append(["4", "El lider directo aprueba el gasto", "3 dias habiles"])
ws2.append(["5", "Finanzas procesa el reembolso", "5 dias habiles"])
ws2.append(["6", "El valor se acredita en la nomina siguiente", ""])
ws3 = wb.create_sheet("Resumen Gastos")
ws3.append(["Tema", "Descripcion"])
ws3.append(["Tope alimentacion", "El tope diario para gastos de alimentacion es de $45,000 pesos. No requiere autorizacion previa. Debe presentar factura."])
ws3.append(["Tope transporte", "Transporte terrestre: hasta $80,000 diarios. Transporte aereo nacional requiere autorizacion del lider."])
ws3.append(["Hospedaje", "Tope diario de $250,000 para hospedaje. Requiere autorizacion del lider directo."])
ws3.append(["Viaticos internacionales", "Para viajes internacionales el tope es de USD 120 por dia e incluye alimentacion y gastos menores. Requiere autorizacion de direccion."])
ws3.append(["Gasolina", "Si se usa vehiculo propio, se reembolsa a $1,800 por kilometro segun registro de kilometraje."])
ws3.append(["Procedimiento reembolso", "El colaborador paga de su bolsillo, sube la factura al sistema en maximo 15 dias, el lider aprueba en 3 dias, y finanzas procesa el reembolso en 5 dias habiles. El valor se acredita en la siguiente nomina."])
wb.save(str(sh / "fin" / "FIN-005.xlsx"))
print("FIN-005.xlsx OK (con hoja Resumen)")

print("\nXLSX mejorados OK")
